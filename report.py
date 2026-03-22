"""
CSV → HTML + Excel Rapor Üretici
"""

import csv
import os
import webbrowser
from datetime import datetime


# ── Yardımcı ──────────────────────────────────────────────────────────────────

def _parse_rows(csv_path: str) -> list[dict]:
    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    if not rows:
        raise ValueError("CSV dosyası boş.")
    return rows


def _summary(rows: list[dict]) -> dict:
    last = rows[-1]
    timestamp_raw = rows[0].get("timestamp", "")
    try:
        session_date = datetime.fromisoformat(timestamp_raw).strftime("%d.%m.%Y %H:%M")
    except Exception:
        session_date = timestamp_raw[:16] if timestamp_raw else "—"

    criterion_trial = next(
        (int(r["trial"]) for r in rows if r.get("criterion_reached", "0") == "1"), None
    )
    return {
        "animal_id":       rows[0].get("animal_id", "—"),
        "session_date":    session_date,
        "total_trials":    len(rows),
        "hit_rate":        float(last.get("hit_rate",  0) or 0),
        "cr_rate":         float(last.get("cr_rate",   0) or 0),
        "d_prime":         float(last.get("d_prime",   0) or 0),
        "rewarded":        int(last.get("rewarded",   0) or 0),
        "punished":        int(last.get("punished",   0) or 0),
        "omission":        int(last.get("omission",   0) or 0),
        "cr_count":        int(last.get("correct_rejection", 0) or 0),
        "total_licks":     sum(int(r.get("lick_count",  0) or 0) for r in rows),
        "total_iti":       sum(int(r.get("iti_presses", 0) or 0) for r in rows),
        "sync_misses":     sum(1 for r in rows if r.get("sound_confirmed", "1") == "0"),
        "criterion_trial": criterion_trial,
    }


def _dprime_color_html(v: float) -> str:
    if v >= 2.0: return "#4ade80"
    if v >= 1.0: return "#fbbf24"
    return "#f87171"


# ── HTML ──────────────────────────────────────────────────────────────────────

RESULT_BG = {
    "rewarded":          "#14532d",
    "punished":          "#431407",
    "omission":          "#450a0a",
    "correct_rejection": "#1e3a5f",
}
RESULT_LABEL = {
    "rewarded":          "Ödül",
    "punished":          "Ceza",
    "omission":          "Omission",
    "correct_rejection": "Correct Rejection",
}


def generate_html(csv_path: str) -> str:
    rows = _parse_rows(csv_path)
    s    = _summary(rows)

    # Trial satırları
    trial_rows_html = ""
    for r in rows:
        result  = r.get("result", "")
        bg      = RESULT_BG.get(result, "#1e1e2e")
        label   = RESULT_LABEL.get(result, result)
        rt_ds   = f"{float(r['rt_from_ds_s']):.3f}"   if r.get("rt_from_ds_s")    else "—"
        rt_lev  = f"{float(r['rt_from_lever_s']):.3f}" if r.get("rt_from_lever_s") else "—"
        licks   = r.get("lick_count",  "0") or "0"
        iti_p   = r.get("iti_presses", "0") or "0"
        hr      = f"{float(r['hit_rate'])*100:.1f}%"  if r.get("hit_rate")  else "—"
        cr      = f"{float(r['cr_rate'])*100:.1f}%"   if r.get("cr_rate")   else "—"
        dp_val  = float(r.get("d_prime", 0) or 0)
        dp_str  = f"{dp_val:.2f}" if r.get("d_prime") else "—"
        dp_col  = _dprime_color_html(dp_val)
        crit    = "✓" if r.get("criterion_reached", "0") == "1" else ""
        sync    = "" if r.get("sound_confirmed", "1") == "1" else "⚠"

        trial_rows_html += f"""
        <tr style="background:{bg}">
            <td>{r.get('trial','')}</td>
            <td>{r.get('ds_type','')}</td>
            <td><b>{label}</b></td>
            <td>{rt_ds}</td>
            <td>{rt_lev}</td>
            <td>{licks}</td>
            <td style="{'color:#fb923c;font-weight:bold' if int(iti_p)>0 else ''}">{iti_p}</td>
            <td>{hr}</td>
            <td>{cr}</td>
            <td style="color:{dp_col};font-weight:bold">{dp_str}</td>
            <td style="color:#4ade80;font-weight:bold">{crit}</td>
            <td style="color:#f87171;font-weight:bold">{sync}</td>
        </tr>"""

    if s["criterion_trial"]:
        crit_badge = f'<span class="badge badge-green">✓ Trial {s["criterion_trial"]}\'de kritere ulaşıldı</span>'
    else:
        crit_badge = '<span class="badge badge-red">✗ Kriter bu oturumda sağlanamadı</span>'

    dp_col_main = _dprime_color_html(s["d_prime"])

    iti_warn  = f'<div class="warn">⚠ Bu oturumda toplam <b>{s["total_iti"]}</b> ITI lever press kaydedildi (dürtüsellik).</div>' if s["total_iti"] > 0 else ""
    sync_warn = f'<div class="warn">⚠ <b>{s["sync_misses"]}</b> trialda Avisoft ses onayı alınamadı (sound_confirmed=0).</div>'    if s["sync_misses"] > 0 else ""

    html = f"""<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<title>Oturum Raporu — {s['animal_id']}</title>
<style>
  :root {{
    --bg:      #0f0f1a;
    --surface: #1a1a2e;
    --card:    #16213e;
    --border:  #2d2d4e;
    --text:    #e2e8f0;
    --muted:   #94a3b8;
    --accent:  #6366f1;
    --blue:    #3b82f6;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: var(--bg); color: var(--text); padding: 36px; }}
  h1   {{ font-size: 26px; color: var(--accent); margin-bottom: 4px; }}
  h2   {{ font-size: 16px; color: var(--blue); margin: 32px 0 12px; border-bottom: 1px solid var(--border); padding-bottom: 6px; }}
  .meta {{ color: var(--muted); font-size: 13px; margin-bottom: 20px; }}
  .meta b {{ color: var(--text); }}
  .badge {{ display:inline-block; padding: 5px 14px; border-radius: 6px; font-size: 13px; font-weight: bold; margin-bottom: 4px; }}
  .badge-green {{ background: #14532d; color: #4ade80; }}
  .badge-red   {{ background: #450a0a; color: #f87171; }}
  .summary-grid {{
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin: 16px 0 24px;
  }}
  .card {{
    background: var(--card); border: 1px solid var(--border); border-radius: 10px;
    padding: 18px; text-align: center;
  }}
  .card .value {{ font-size: 30px; font-weight: bold; color: var(--accent); }}
  .card .label {{ font-size: 11px; color: var(--muted); margin-top: 6px; }}
  .outcome-grid {{
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 12px 0;
  }}
  .oc {{ border-radius: 10px; padding: 14px; text-align: center; }}
  .oc .num {{ font-size: 26px; font-weight: bold; }}
  .oc .lbl {{ font-size: 11px; margin-top: 4px; opacity: 0.85; }}
  .warn {{
    background: #292207; border: 1px solid #854d0e; border-radius: 6px;
    padding: 10px 16px; margin: 10px 0; font-size: 13px; color: #fbbf24;
  }}
  table {{
    width: 100%; border-collapse: collapse; font-size: 12.5px;
    background: var(--surface); border-radius: 10px; overflow: hidden;
  }}
  th {{
    background: var(--accent); color: white; padding: 10px 8px;
    text-align: center; font-size: 11px; font-weight: 600;
  }}
  td {{ padding: 7px 8px; text-align: center; border-bottom: 1px solid var(--border); color: var(--text); }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover {{ filter: brightness(1.15); }}
  .legend {{ display: flex; gap: 16px; flex-wrap: wrap; margin: 10px 0; font-size: 12px; color: var(--muted); }}
  .legend-item {{ display: flex; align-items: center; gap: 6px; }}
  .legend-dot {{ width: 12px; height: 12px; border-radius: 3px; }}
  .footer {{ margin-top: 40px; font-size: 11px; color: var(--border); text-align: center; }}
  @media print {{
    body {{ background: white; color: black; padding: 16px; }}
    .card {{ border: 1px solid #ccc; }}
  }}
</style>
</head>
<body>

<h1>Oturum Raporu</h1>
<div class="meta">
  <b>Hayvan ID:</b> {s['animal_id']} &nbsp;|&nbsp;
  <b>Tarih:</b> {s['session_date']} &nbsp;|&nbsp;
  <b>Toplam Trial:</b> {s['total_trials']} &nbsp;|&nbsp;
  <b>CSV:</b> {os.path.basename(csv_path)}
</div>

{crit_badge}

<h2>Diskriminasyon Metrikleri</h2>
<div class="summary-grid">
  <div class="card">
    <div class="value">{s['hit_rate']*100:.1f}%</div>
    <div class="label">Hit Rate (DS+)</div>
  </div>
  <div class="card">
    <div class="value">{s['cr_rate']*100:.1f}%</div>
    <div class="label">Correct Rejection (DS−)</div>
  </div>
  <div class="card">
    <div class="value" style="color:{dp_col_main}">{s['d_prime']:.2f}</div>
    <div class="label">d' (d-prime)</div>
  </div>
  <div class="card">
    <div class="value">{s['total_licks']}</div>
    <div class="label">Toplam Lick</div>
  </div>
</div>

<h2>Trial Sonuçları</h2>
<div class="outcome-grid">
  <div class="oc" style="background:#14532d;color:#4ade80">
    <div class="num">{s['rewarded']}</div><div class="lbl">Ödül (Rewarded)</div>
  </div>
  <div class="oc" style="background:#431407;color:#fb923c">
    <div class="num">{s['punished']}</div><div class="lbl">Ceza (Punished)</div>
  </div>
  <div class="oc" style="background:#450a0a;color:#f87171">
    <div class="num">{s['omission']}</div><div class="lbl">Omission (DS+↓)</div>
  </div>
  <div class="oc" style="background:#1e3a5f;color:#93c5fd">
    <div class="num">{s['cr_count']}</div><div class="lbl">Correct Rejection (DS−↓)</div>
  </div>
</div>

{iti_warn}{sync_warn}

<h2>Trial Detayları</h2>
<div class="legend">
  <div class="legend-item"><div class="legend-dot" style="background:#14532d"></div>Ödül</div>
  <div class="legend-item"><div class="legend-dot" style="background:#431407"></div>Ceza</div>
  <div class="legend-item"><div class="legend-dot" style="background:#450a0a"></div>Omission</div>
  <div class="legend-item"><div class="legend-dot" style="background:#1e3a5f"></div>Correct Rejection</div>
  <div class="legend-item" style="color:#4ade80">✓ = Kriter sağlandı</div>
  <div class="legend-item" style="color:#f87171">⚠ = Ses onayı yok</div>
</div>

<table>
  <thead>
    <tr>
      <th>Trial</th><th>DS</th><th>Sonuç</th>
      <th>RT (DS) sn</th><th>RT (Lever) sn</th>
      <th>Lick</th><th>ITI Press</th>
      <th>Hit Rate</th><th>CR Rate</th><th>d'</th>
      <th>Kriter</th><th>Ses</th>
    </tr>
  </thead>
  <tbody>{trial_rows_html}</tbody>
</table>

<div class="footer">
  Boğaziçi Üniversitesi Davranışsal Nörobilim Laboratuvarı —
  Oluşturulma: {datetime.now().strftime("%d.%m.%Y %H:%M")}
</div>
</body>
</html>"""

    report_path = csv_path.replace(".csv", "_rapor.html")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(html)
    return report_path


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_excel(csv_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    rows = _parse_rows(csv_path)
    s    = _summary(rows)
    wb   = Workbook()

    # ── Renk paletleri ────────────────────────────────────────────────────────
    C = {
        "bg":       "0F0F1A", "surface":  "1A1A2E", "accent":   "6366F1",
        "blue":     "3B82F6", "text":     "E2E8F0", "muted":    "94A3B8",
        "green_bg": "14532D", "green_fg": "4ADE80",
        "red_bg":   "450A0A", "red_fg":   "F87171",
        "orange_bg":"431407", "orange_fg":"FB923C",
        "navy_bg":  "1E3A5F", "navy_fg":  "93C5FD",
        "warn_bg":  "292207", "warn_fg":  "FBBF24",
        "header":   "6366F1",
    }

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(hex_color, bold=False, size=11):
        return Font(color=hex_color, bold=bold, size=size, name="Segoe UI")

    thin = Side(style="thin", color="2D2D4E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ════════════════════════════════════════════════════════════════════════
    # SAYFA 1 — ÖZET
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Özet"
    ws1.sheet_view.showGridLines = False
    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    def write(ws, row, col, val, bg=None, fg="E2E8F0", bold=False, size=11, align=True):
        cell = ws.cell(row=row, column=col, value=val)
        if bg:
            cell.fill = fill(bg)
        cell.font  = font(fg, bold=bold, size=size)
        cell.border = border
        if align:
            cell.alignment = center
        return cell

    # Başlık
    ws1.merge_cells("A1:F1")
    c = ws1.cell(row=1, column=1, value=f"OTURUM RAPORU — {s['animal_id']}")
    c.fill = fill(C["accent"]); c.font = Font(color="FFFFFF", bold=True, size=16, name="Segoe UI")
    c.alignment = center; ws1.row_dimensions[1].height = 36

    # Meta
    ws1.merge_cells("A2:F2")
    meta_val = f"Tarih: {s['session_date']}   |   Toplam Trial: {s['total_trials']}   |   CSV: {os.path.basename(csv_path)}"
    c = ws1.cell(row=2, column=1, value=meta_val)
    c.fill = fill(C["surface"]); c.font = font(C["muted"], size=10)
    c.alignment = center; ws1.row_dimensions[2].height = 20

    # Kriter
    ws1.merge_cells("A3:F3")
    if s["criterion_trial"]:
        crit_text = f"✓  Trial {s['criterion_trial']}'de kritere ulaşıldı"
        crit_bg, crit_fg = C["green_bg"], C["green_fg"]
    else:
        crit_text = "✗  Kriter bu oturumda sağlanamadı"
        crit_bg, crit_fg = C["red_bg"], C["red_fg"]
    c = ws1.cell(row=3, column=1, value=crit_text)
    c.fill = fill(crit_bg); c.font = Font(color=crit_fg, bold=True, size=12, name="Segoe UI")
    c.alignment = center; ws1.row_dimensions[3].height = 28

    ws1.row_dimensions[4].height = 10  # boşluk

    # Metrik kartları — başlıklar
    headers = ["Hit Rate (DS+)", "Correct Rejection (DS−)", "d' (d-prime)", "Toplam Lick", "Toplam ITI Press", "Sync Miss"]
    for i, h in enumerate(headers, 1):
        write(ws1, 5, i, h, bg=C["blue"], fg="FFFFFF", bold=True, size=10)
    ws1.row_dimensions[5].height = 22

    # Metrik kartları — değerler
    dp_fg = C["green_fg"] if s["d_prime"] >= 2 else (C["warn_fg"] if s["d_prime"] >= 1 else C["red_fg"])
    values = [
        (f"{s['hit_rate']*100:.1f}%",  C["green_fg"]),
        (f"{s['cr_rate']*100:.1f}%",   C["green_fg"]),
        (f"{s['d_prime']:.2f}",        dp_fg),
        (str(s["total_licks"]),        C["text"]),
        (str(s["total_iti"]),          C["warn_fg"] if s["total_iti"] > 0 else C["text"]),
        (str(s["sync_misses"]),        C["red_fg"]  if s["sync_misses"] > 0 else C["text"]),
    ]
    for i, (val, fg) in enumerate(values, 1):
        write(ws1, 6, i, val, bg=C["surface"], fg=fg, bold=True, size=16)
    ws1.row_dimensions[6].height = 40

    ws1.row_dimensions[7].height = 10  # boşluk

    # Outcome tablosu
    out_headers = ["Ödül (Rewarded)", "Ceza (Punished)", "Omission (DS+↓)", "Correct Rejection (DS−↓)"]
    out_values  = [s["rewarded"], s["punished"], s["omission"], s["cr_count"]]
    out_bgs     = [C["green_bg"], C["orange_bg"], C["red_bg"], C["navy_bg"]]
    out_fgs     = [C["green_fg"], C["orange_fg"], C["red_fg"],  C["navy_fg"]]

    for i, (h, v, bg, fg) in enumerate(zip(out_headers, out_values, out_bgs, out_fgs), 1):
        write(ws1, 8, i, h,  bg=bg, fg=fg, bold=True, size=10)
        write(ws1, 9, i, v,  bg=bg, fg=fg, bold=True, size=20)
    ws1.row_dimensions[8].height = 22
    ws1.row_dimensions[9].height = 40

    # ════════════════════════════════════════════════════════════════════════
    # SAYFA 2 — TRIAL DETAYLARI
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Trial Detayları")
    ws2.sheet_view.showGridLines = False

    col_widths = [8, 8, 20, 14, 14, 8, 10, 10, 10, 8, 8, 8]
    col_headers = [
        "Trial", "DS", "Sonuç", "RT (DS) sn", "RT (Lever) sn",
        "Lick", "ITI Press", "Hit Rate", "CR Rate", "d'", "Kriter", "Ses"
    ]
    for i, (w, h) in enumerate(zip(col_widths, col_headers), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        write(ws2, 1, i, h, bg=C["header"], fg="FFFFFF", bold=True, size=10)
    ws2.row_dimensions[1].height = 24

    result_style = {
        "rewarded":          (C["green_bg"],  C["green_fg"]),
        "punished":          (C["orange_bg"], C["orange_fg"]),
        "omission":          (C["red_bg"],    C["red_fg"]),
        "correct_rejection": (C["navy_bg"],   C["navy_fg"]),
    }

    for rx, r in enumerate(rows, 2):
        result = r.get("result", "")
        rbg, rfg = result_style.get(result, (C["surface"], C["text"]))
        dp_val = float(r.get("d_prime", 0) or 0)
        dp_fg  = C["green_fg"] if dp_val >= 2 else (C["warn_fg"] if dp_val >= 1 else C["red_fg"])

        def w(col, val, fg=C["text"]):
            write(ws2, rx, col, val, bg=rbg, fg=fg)

        w(1,  int(r.get("trial", 0) or 0))
        w(2,  r.get("ds_type", ""))
        write(ws2, rx, 3, RESULT_LABEL.get(result, result), bg=rbg, fg=rfg, bold=True)
        w(4,  float(r["rt_from_ds_s"])    if r.get("rt_from_ds_s")    else "—")
        w(5,  float(r["rt_from_lever_s"]) if r.get("rt_from_lever_s") else "—")
        w(6,  int(r.get("lick_count",  0) or 0))
        iti = int(r.get("iti_presses", 0) or 0)
        write(ws2, rx, 7, iti, bg=rbg, fg=C["warn_fg"] if iti > 0 else C["text"], bold=iti > 0)
        w(8,  f"{float(r['hit_rate'])*100:.1f}%" if r.get("hit_rate") else "—")
        w(9,  f"{float(r['cr_rate'])*100:.1f}%"  if r.get("cr_rate")  else "—")
        write(ws2, rx, 10, f"{dp_val:.2f}" if r.get("d_prime") else "—", bg=rbg, fg=dp_fg, bold=True)
        crit = r.get("criterion_reached", "0") == "1"
        write(ws2, rx, 11, "✓" if crit else "", bg=rbg, fg=C["green_fg"], bold=True)
        sync_ok = r.get("sound_confirmed", "1") == "1"
        write(ws2, rx, 12, "" if sync_ok else "⚠", bg=rbg, fg=C["red_fg"], bold=True)
        ws2.row_dimensions[rx].height = 18

    # ── Kaydet ────────────────────────────────────────────────────────────────
    excel_path = csv_path.replace(".csv", "_rapor.xlsx")
    wb.save(excel_path)
    return excel_path


# ── Her ikisini birden üret ────────────────────────────────────────────────────

def generate_report(csv_path: str) -> tuple[str, str]:
    """HTML ve Excel raporu üretir. (html_path, excel_path) döner."""
    return generate_html(csv_path), generate_excel(csv_path)


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Kullanım: python3 report.py logs/session_RAT_01_....csv")
        sys.exit(1)
    html_p, xlsx_p = generate_report(sys.argv[1])
    print(f"HTML:  {html_p}")
    print(f"Excel: {xlsx_p}")
    webbrowser.open(f"file://{os.path.abspath(html_p)}")
